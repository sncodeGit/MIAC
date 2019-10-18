# -*- coding: utf-8 -*-
"""
Created on Sat Oct 19 01:50:24 2019
Вывод id МО, имеющих один код типа МО
Id задаётся в константе SEARCH_CODE

Необходимо:
файл manual_MO.xlsx, выгрзку в excel из справочника МО
файл num_of_hosp_records формата id_МО:количество_записей_в_выгрузке

@author: inter000
"""

import openpyxl
from openpyxl.utils import column_index_from_string

MAX_ROW = 1834 # Количество строк в manual_MO
SEARCH_CODE = 10487 # Код типа больницы, по которому ведётся поиск МО

file = open('num_of_hosp_records', 'r')
lines = file.readlines()

MO_class = {} # key - код МО, value - код типа больницы
for line in lines:
    MO_class[line.split(':')[0]] = -1 # Split: MO_code:count_of_rows

file.close()

wb = openpyxl.load_workbook('manual_MO.xlsx') # Открываем файл xlsx
main_sheet = wb['Reference data'] # Выбираем таблицу файла

for i in range(2, MAX_ROW + 1):
    # AJ - код столбца 'Код' (код МО)
    MO_ind = column_index_from_string('AJ')
    # Y - код столбца 'medorgtype, codes' 
    org_ind = column_index_from_string('Y')
    # AE - код столбца 'medobjtype, codes' 
    obj_ind = column_index_from_string('AE')
    
    MO_value = main_sheet.cell(row = i, column = MO_ind).value
    if MO_value in MO_class:
        org_value = main_sheet.cell(row = i, column = org_ind).value
        obj_value = main_sheet.cell(row = i, column = obj_ind).value
        if org_value is not None:
            MO_class[MO_value] = int(org_value)
        elif obj_value is not None:
            MO_class[MO_value] = int(obj_value)
            
for MO_code in MO_class.keys():
    if MO_class[MO_code] == SEARCH_CODE:
        print(MO_code)
